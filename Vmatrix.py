import os
import pandas as pd
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.comments import Comment


base_root = os.getcwd()

# 1. 필요한 스킬 목록 구성
필요한_스킬 = ['메익', '암살', '쿼드', '마크', '플레어', '배오새']

# 2. 보유 V 코어 목록 구성
v_코어 = {
    'V1': ['메익', '암살', '쿼드'],
    'V2': ['암살', '메익', '쿼드'],
    'V3': ['플레어', '마크', '배오새'],
    'V4': ['메익', '암살', '마크'],
    'V5': ['배오새', '마크', '플레어'],
    'V6': ['메익', '배오새', '쿼드'],
    'V7': ['메익', '쿼드', '암살']
    # ... 여러개의 V 코어 추가
}

# 3. 필요한 스킬을 2번씩 중첩하여 강화하는 조합 찾기

v_코어_수 = 4  # 장착할 v 코어 수  

combinations = list(combinations(v_코어.keys(), v_코어_수))  # V 코어를 조합

v_core_values_list = []

for combo in combinations:
    v_core_values = [v_코어[v_core] for v_core in combo]  # 각 조합에 대한 V 코어의 value 값을 추출
    main_skills = [v_core_value[0] for v_core_value in v_core_values]  # 각 조합에서 맨 앞의 스킬을 추출
    if len(set(main_skills)) == v_코어_수:  # 맨 앞의 스킬이 모두 다르면 추가
        v_core_data = {f'V{i}': [] for i in range(1, v_코어_수 + 1)}
        for value in v_core_values:
            for i in range(1, v_코어_수 + 1):
                if i <= len(value):
                    v_core_data[f'V{i}'].append(value[i - 1])
                else:
                    v_core_data[f'V{i}'].append(None)
        v_core_values_list.append(pd.DataFrame(v_core_data))

# 결과를 엑셀 파일로 저장
if v_core_values_list:
    output_directory = base_root  # Change 'output' to your desired directory name
    os.makedirs(output_directory, exist_ok=True)  # Create the output directory if it doesn't exist

    excel_path = os.path.join(output_directory, '추천목록.xlsx')
    with pd.ExcelWriter(excel_path, engine='xlsxwriter', mode='w') as writer:
        for idx, df_combination in enumerate(v_core_values_list):
            df_combination.to_excel(writer, sheet_name=f'추천 조합 {idx + 1}', index=False, header=False)

    # Add comments to cell A1 in the sheets
    workbook = load_workbook(excel_path)
    for idx, sheet_name in enumerate(workbook.sheetnames):
        worksheet = workbook[sheet_name]
        
    workbook.save(excel_path)

    print(f'엑셀 파일이 생성되었습니다. 경로: {excel_path}')
else:
    print('조합이 없습니다. V 코어 또는 필요한 스킬을 확인하세요.')


# 사용자 입력 받기
print("1. 필요한 스킬 목록 작성:")
필요한_스킬 = input("필요한 스킬을 띄어쓰기로 구분하여 입력하세요 (예: 메익 암살 쿼드): ").split()

print("\n2. V 코어 목록 작성:")
print("예시 V 코어 목록:")
print("V1: 메익 암살 쿼드")
print("V2: 암살 메익 쿼드")
# ... 기타 V 코어 목록 예시 출력

v_코어 = {}
v_core_index = 1  # 초기 인덱스 설정
while True:
    v_core_skills = input(f"V{v_core_index} 코어의 스킬을 띄어쓰기로 구분하여 입력하세요 (종료하려면 'exit' 입력): ").split()
    
    if v_core_skills[0].lower() == 'exit':
        break
    
    if all(skill.replace('_', '').isalpha() or skill.replace('_', '').isalpha() and skill.replace('_', '').isascii() for skill in v_core_skills):
        v_코어[f'V{v_core_index}'] = v_core_skills
        v_core_index += 1
    else:
        print("스킬은 알파벳 또는 한글로만 입력하세요.")

# 추천 조합 찾기
v_코어_수 = int(input("\n3. V 코어 수를 입력하세요: "))
combinations = list(combinations(v_코어.keys(), v_코어_수))
valid_combinations = []

for combo in combinations:
    main_skills = [v_코어[v_core][0] for v_core in combo]
    if len(set(main_skills)) == v_코어_수:
        valid_combinations.append(combo)

# 결과 디스플레이
print("\n추천 조합 목록:")
for idx, combo in enumerate(valid_combinations, start=1):
    print(f"{idx}. {', '.join(combo)}")

